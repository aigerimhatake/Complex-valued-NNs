import logging
import cvnn
import cvnn.layers as layers
import cvnn.dataset as dp
from cvnn.dataset import Dataset
from cvnn.cvnn_model import CvnnModel
from cvnn.data_analysis import MonteCarloAnalyzer, confusion_matrix, SeveralMonteCarloComparison
from cvnn.layers import Dense
from cvnn.utils import create_folder, transform_to_real, randomize
import tensorflow as tf
import pandas as pd
import copy
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from tqdm import tqdm
import numpy as np
from pdb import set_trace
from time import sleep

logger = logging.getLogger(cvnn.__name__)


class MonteCarlo:

    def __init__(self):
        self.models = []
        self.pandas_full_data = pd.DataFrame()
        self.confusion_matrix = []
        self.monte_carlo_analyzer = MonteCarloAnalyzer()  # All at None

    def add_model(self, model):
        self.models.append(model)

    def run(self, x, y, data_summary='', polar=False, do_conf_mat=True, ratio=0.8,
            iterations=100, learning_rate=0.01, epochs=10, batch_size=100,
            shuffle=False, debug=False, display_freq=160, checkpoints=False):
        x, y = randomize(x, y)
        # Reset data frame
        self.pandas_full_data = pd.DataFrame()
        if do_conf_mat:
            for i in range(len(self.models)):
                self.confusion_matrix.append({"name": "model_name", "matrix": pd.DataFrame()})
        self.save_summary_of_run(self._run_summary(iterations, learning_rate, epochs, batch_size, shuffle),
                                 data_summary)
        if not debug:
            pbar = tqdm(total=iterations)
        for it in range(iterations):
            if debug:
                logger.info("Iteration {}/{}".format(it + 1, iterations))
            else:
                pbar.update()
            if shuffle:  # shuffle all data at each iteration
                x, y = randomize(x, y)
            for i, model in enumerate(self.models):
                if model.is_complex():
                    x_fit = x
                else:
                    x_fit = transform_to_real(x, polar=polar)
                test_model = copy.deepcopy(model)
                test_model.fit(x_fit, y, ratio=ratio,
                               learning_rate=learning_rate, epochs=epochs, batch_size=batch_size,
                               verbose=debug, fast_mode=True, save_txt_fit_summary=False, display_freq=display_freq,
                               save_csv_history=True)
                self.pandas_full_data = pd.concat([self.pandas_full_data,
                                                   test_model.plotter.get_full_pandas_dataframe()], sort=False)
                if do_conf_mat:
                    dataset = dp.Dataset(x_fit, y, ratio=ratio)
                    self.confusion_matrix[i]["name"] = test_model.name
                    self.confusion_matrix[i]["matrix"] = pd.concat((self.confusion_matrix[i]["matrix"],
                                                                    test_model.get_confusion_matrix(dataset.x_test,
                                                                                                    dataset.y_test)))
            if checkpoints:
                # Save checkpoint in case Monte Carlo stops in the middle
                self.pandas_full_data.to_csv(self.monte_carlo_analyzer.path / "run_data.csv", index=False)
        if not debug:
            pbar.close()
        self.pandas_full_data = self.pandas_full_data.reset_index(drop=True)
        conf_mat = None
        if do_conf_mat:
            conf_mat = self.confusion_matrix
        self.monte_carlo_analyzer.set_df(self.pandas_full_data, conf_mat)

    @staticmethod
    def _run_summary(iterations, learning_rate, epochs, batch_size, shuffle):
        ret_str = "Monte Carlo run\n"
        ret_str += "\tIterations: {}\n".format(iterations)
        ret_str += "\tepochs: {}\n".format(epochs)
        ret_str += "\tbatch_size: {}\n".format(batch_size)
        ret_str += "\tLearning Rate: {}\n".format(learning_rate)
        if shuffle:
            ret_str += "\tShuffle data at each iteration\n"
        else:
            ret_str += "\tData is not shuffled at each iteration\n"
        return ret_str

    def save_summary_of_run(self, run_summary, data_summary):
        with open(str(self.monte_carlo_analyzer.path / "run_summary.txt"), "w") as file:
            file.write(run_summary)
            file.write(data_summary)
            file.write("Models:\n")
            for model in self.models:
                file.write(model.summary())


class RealVsComplex(MonteCarlo):

    def __init__(self, complex_model):
        super().__init__()
        # add models
        self.add_model(complex_model)
        self.add_model(complex_model.get_real_equivalent(name="real_network"))


def run_montecarlo(iterations=1000, m=10000, n=128, param_list=None, open_dataset=None,
                   epochs=150, batch_size=100, display_freq=None, learning_rate=0.01,
                   shape_raw=None, activation='cart_relu', debug=False, polar=False, do_all=True, dropout=None):
    # Get parameters
    if shape_raw is None:
        shape_raw = [64]
    if open_dataset:
        if param_list is not None:
            logger.error("If the parameter to open_dataset is passed, giving param_list makes no sense")
            sys.exit(-1)
        if not m == 10000:
            logger.error("If the parameter to open_dataset is passed, giving m makes no sense")
            sys.exit(-1)
        if not n == 128:
            logger.error("If the parameter to open_dataset is passed, giving n makes no sense")
            sys.exit(-1)
        dataset = dp.OpenDataset(open_dataset)
    else:
        if param_list is None:
            param_list = [
                [0.5, 1, 1],
                [-0.5, 1, 1]
            ]
        dataset = dp.CorrelatedGaussianCoeffCorrel(m, n, param_list, debug=False)
    if display_freq is None:
        display_freq = int(m * dataset.num_classes * 0.8 / batch_size)

    # Create complex network
    input_size = dataset.x.shape[1]  # Size of input
    output_size = dataset.y.shape[1]  # Size of output
    if not len(shape_raw) > 0:
        logger.error("Shape raw was empty")
        sys.exit(-1)
    layers.ComplexLayer.last_layer_output_dtype = None
    layers.ComplexLayer.last_layer_output_size = None
    shape = [Dense(input_size=input_size, output_size=shape_raw[0], activation=activation,
                   input_dtype=np.complex64, dropout=dropout)]
    for i in range(1, len(shape_raw)):
        shape.append(Dense(output_size=shape_raw[i], activation=activation, dropout=None))
    shape.append(Dense(output_size=output_size, activation='softmax_real'))

    complex_network = CvnnModel(name="complex_network", shape=shape, loss_fun=tf.keras.losses.categorical_crossentropy,
                                verbose=False, tensorboard=False)

    # Monte Carlo
    monte_carlo = RealVsComplex(complex_network)
    if not open_dataset:
        dataset.save_data(monte_carlo.monte_carlo_analyzer.path)
    sleep(1)    # I have error if not because not enough time passed since creation of models to be in diff folders
    monte_carlo.run(dataset.x, dataset.y, iterations=iterations, learning_rate=learning_rate,
                    epochs=epochs, batch_size=batch_size, display_freq=display_freq,
                    shuffle=False, debug=debug, data_summary=dataset.summary(), polar=polar)
    if do_all:
        monte_carlo.monte_carlo_analyzer.do_all()

    # TODO: Separate this part of code PLEASE!
    filename = './log/mlp_monte_carlo_summary.xlsx'
    fieldnames = ['path', 'HL', 'Shape', 'Dropout', '# Classes', "Polar Mode", "Learning Rate",
                  "Activation Function", "Dataset Size", 'Feature Size', 'epochs', 'batch size',
                  "Winner", "CVNN median", "RVNN median", 'CVNN IQR', 'RVNN IQR', "Comments"]
    max_epoch = monte_carlo.pandas_full_data['epoch'].max()
    epoch_filter = monte_carlo.pandas_full_data['epoch'] == max_epoch
    complex_filter = monte_carlo.pandas_full_data['network'] == "complex network"
    real_filter = monte_carlo.pandas_full_data['network'] == "real network"
    complex_last_epochs = monte_carlo.pandas_full_data[epoch_filter & complex_filter]
    real_last_epochs = monte_carlo.pandas_full_data[epoch_filter & real_filter]
    complex_median = complex_last_epochs['test accuracy'].median()
    real_median = real_last_epochs['test accuracy'].median()
    row_data = [str(monte_carlo.monte_carlo_analyzer.path), str(len(shape_raw)), str(shape_raw),
                str(dropout), str(dataset.y.shape[1]),
                'Yes' if polar == 'Apple' else 'No', learning_rate, activation, str(dataset.x.shape[0]),
                str(dataset.x.shape[1]), epochs, batch_size,
                'CVNN' if complex_median > real_median else 'RVNN',
                complex_median, real_median,
                complex_last_epochs['test accuracy'].quantile(.75)
                - complex_last_epochs['test accuracy'].quantile(.25),
                real_last_epochs['test accuracy'].quantile(.75)
                - real_last_epochs['test accuracy'].quantile(.25),
                ]
    file_exists = os.path.isfile(filename)
    if file_exists:
        wb = load_workbook(filename)
        ws = wb.worksheets[0]
    else:
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.append(fieldnames)
    ws.append(row_data)
    tab = Table(displayName="Table1", ref="A1:R" + str(ws.max_row))
    percentage_cols = ['N', 'O', 'P', 'Q']
    for col in percentage_cols:
        ws[col + str(ws.max_row)].number_format = '0.00%'
    ws.add_table(tab)
    wb.save(filename)

    return str(monte_carlo.monte_carlo_analyzer.path / "run_data.csv")


if __name__ == "__main__":
    # Base case with one hidden layer size 64 and dropout 0.5
    run_montecarlo(iterations=1, epochs=10, do_all=True, open_dataset="./data/MLSP/")
